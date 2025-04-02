import pandas as pd
import streamlit as st
from io import BytesIO
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side

def process_data(route_data, products_data):
    # Process Route Data
    route_df = pd.read_csv(route_data)
    
    # Handle column names with potential whitespace variations
    route_df.columns = route_df.columns.str.strip()
    
    # Column cleaning and renaming
    route_df = route_df.rename(columns={
        'Note (Order)': 'Note',
        'Drive time (minutes)': 'time',
        'Distance (km)': 'dist'
    })
    
    # Forward fill missing values
    fill_cols = ['Route', 'Driver', 'Stop', 'Address', 'Shipping name']
    route_df[fill_cols] = route_df[fill_cols].ffill()
    
    # Address processing
    route_df['Address'] = route_df['Address'].str.split(',').str[0].ffill()
    
    # Process Products Data
    df_plist = pd.read_csv(products_data)
    df_plist = df_plist[['Title', 'Tags', 'Variant SKU', 'Vendor']].dropna(subset=['Tags'])
    df_plist['Product type'] = df_plist['Tags'].apply(determine_product_type)
    df_plist['Lineitem name'] = df_plist['Title']
    
    # Extract item information
    route_df['Lineitem name'] = route_df['Items'].str[3:]
    route_df['Item Count'] = route_df['Items'].str[:3].str.extract('(\d+)')
    
    # Merge product information
    route_df = pd.merge(route_df, df_plist, on='Lineitem name', how='left')
    
    # Create Count column
    route_df['Count'] = np.where(route_df['Item Count'] != "1", route_df['Item Count'], '')
    
    # Select final columns
    final_cols = ['Driver','Stop','Shipping name','Address','Count', 
                 "Item Count", "Lineitem name",'Total items','Note','time','dist']
    
    # Verify columns exist before selection
    missing_cols = [col for col in final_cols if col not in route_df.columns]
    if missing_cols:
        raise ValueError(f"Missing columns in route data: {missing_cols}")
    
    return route_df[final_cols], df_plist

def determine_product_type(tag):
    tag = str(tag).lower()
    if '蔬菜水果类' in tag or '独家订阅包' in tag:
        return 'vege & fruit'
    elif '冰冻冷藏类' in tag:
        return 'cool'
    elif '常温类' in tag:
        return 'general'
    return 'other'

def create_routing_file(route_df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        for driver, group in route_df.groupby('Driver'):
            final = group[['Address', "Count", "Lineitem name", 'Total items', 
                         'Note', 'time', 'dist', 'Stop']]
            
            # Add totals
            sums = final[['Total items', 'time', 'dist']].sum()
            final = final.append(sums.rename('Total'))
            
            final.to_excel(writer, sheet_name=driver, index=False)
            worksheet = writer.sheets[driver]
            worksheet.set_footer(f'&C{driver} &P &20 &"Arial,Bold Italic"')
    
    return format_excel_file(output)

def create_packing_file(packing_df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        for driver, group in packing_df.groupby('Driver'):
            final = group[["Product type", "Count_New", "Lineitem name", 
                         "Count", "Variant SKU", 'Vendor']]
            
            # Add totals
            sums = final[['Count']].sum()
            final = final.append(sums.rename('Total'))
            
            final.to_excel(writer, sheet_name=driver, index=False)
            worksheet = writer.sheets[driver]
            worksheet.set_footer(f'&C{driver} &P &20 &"Arial,Bold Italic"')
    
    return format_excel_file(output)

def format_excel_file(buffer):
    buffer.seek(0)
    wb = load_workbook(buffer)
    for ws in wb.worksheets:
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
        
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = thin_border
        
        ws.freeze_panes = 'A2'
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

# Streamlit UI
st.title("Route and Packing List Generator")

route_file = st.file_uploader("Upload Route Data (CSV)", type="csv")
products_file = st.file_uploader("Upload Products Data (CSV)", type="csv")

if route_file and products_file:
    try:
        route_df, df_plist = process_data(route_file, products_file)
        
        # Generate Packing DataFrame
        packing_df = route_df[['Driver', "Lineitem name", "Item Count"]]
        packing_df['Count'] = packing_df['Item Count'].astype(int)
        packing_df = packing_df.groupby(["Driver", "Lineitem name"]).sum().reset_index()
        packing_df['Count_New'] = np.where(packing_df['Count'] > 1, packing_df['Count'], '')
        packing_df = pd.merge(packing_df, df_plist, on='Lineitem name', how='left')
        packing_df = packing_df.sort_values(by=["Driver", "Product type", "Variant SKU", "Lineitem name"])
        
        # Create files
        routing_file = create_routing_file(route_df)
        packing_file = create_packing_file(packing_df)
        
        # Download buttons
        st.download_button("Download Tuesday Routing", 
                          data=routing_file, 
                          file_name="Tuesday_Routing.xlsx")
        
        st.download_button("Download Tuesday Packing", 
                          data=packing_file, 
                          file_name="Tuesday_Packing.xlsx")
        
    except Exception as e:
        st.error(f"Error processing files: {str(e)}")
